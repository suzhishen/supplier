from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import json
from odoo.tools import groupby
import concurrent.futures
import base64
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl
from io import BytesIO


class FastDataProductTemplateDownload(models.Model):
    _name = "fast.data.product.template.download"
    _description = '数据模板下载'

    name = fields.Char('名称', default='数据模板下载')

    # 服务器Action，打开【数据模板下载】窗口action
    def open_fast_data_product_template_download_action(self):
        datas = self.sudo().search([], limit=1)
        if not datas:
            datas = self.sudo().create({'name': '数据模板下载'})
        module_name = 'fast_data_product_template_download'
        action = self.env.ref(f'{module_name}.fast_data_product_template_download_action').sudo().read()[0]
        form_id = self.env.ref(f'{module_name}.fast_data_product_template_download_form', False)
        action['views'] = [(form_id and form_id.id or False, 'form')]
        action['res_id'] = datas.id
        return action

    # 获取模板下载地址
    def __get_or_create_work_book_xlsx_datas__(self, file_name, headers, row_datas):
        """
        file_name：文件名称 str
        headers：表头 []
        row_datas：数据行 []
        """
        with concurrent.futures.ThreadPoolExecutor() as executor:
            try:
                att_id = self.env['ir.attachment'].sudo().search(
                    [('name', '=', file_name), ('datas', '!=', None), ('res_model', '=', self._name)],
                    order='id desc').filtered(lambda x: x.datas)
                # 如果存在，直接删除附件，防止模板内容有更新，下载的不是最新模板
                att_id.unlink()
                # 创建一个新的工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
                # 设置表头的样式
                header_font = Font(bold=True)
                header_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                header_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                       top=Side(style='medium'),
                                       bottom=Side(style='medium'))
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = header_border
                # 设置数据行的样式
                data_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                # 添加数据行
                for row in row_datas:
                    ws.append(row)
                # 设置数据行的边框样式
                data_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = data_border
                # 调整列宽
                for column in ws.columns:
                    max_length = 7
                    column_name = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_name].width = adjusted_width
                # 保存工作簿到文件
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                # 将 BytesIO 对象转换为 base64 编码的字符串
                xls_binary = base64.b64encode(output.read())
                att_id = self.env['ir.attachment'].create({
                    'name': file_name,
                    'datas': xls_binary,
                    'res_id': self.id,
                    'res_model': self._name,
                })
                base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
                url = f"{base_url}/web/content/ir.attachment/{att_id.id}/datas?download=true"
                return {
                    'type': 'ir.actions.act_url',
                    'target': 'self',
                    'url': url,
                }
            except Exception as e:
                return {
                    'type': 'err',
                    'msg': '发生错误，信息如下：%s！' % str(e)
                }

    # 颜色信息模板
    def btn_blank_template_download(self):
        file_name = '面辅料上传颜色信息模板.xlsx'
        headers = ['po', '加工厂', '装箱单号', '送货日期', '款号', '颜色', '款色', '变体', '数量']
        row_datas = [
            ['F1939-3', '童心缘', '', '2024-05-14', 'M30090', 'BLK', 'M30090-BLK', 'M30090-BLK-S', 10],
            ['F1939-3', '童心缘', '', '2024-05-14', 'M30090', 'BLK', 'M30090-BLK', 'M30090-BLK-S', 20],
            ['F1939-3', '童心缘', '', '2024-05-14', 'M30090', 'BLK', 'M30090-BLK', 'M30090-BLK-M', 30],
        ]
        return self.__get_or_create_work_book_xlsx_datas__(file_name, headers, row_datas)

    # todo 待使用
    def btn_template_download_9(self):
        a = 1

    # todo 待使用
    def btn_template_download_10(self):
        a = 1
